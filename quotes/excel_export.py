from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill(start_color='1c77c3', end_color='1c77c3', fill_type='solid')
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
CELL_BORDER = Border(left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'), top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))
NUMBER_FORMAT = '#,##0.00'
DATE_FORMAT = 'DD/MM/YYYY'

def _apply_header_style(cell):
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = HEADER_ALIGNMENT
    cell.border = CELL_BORDER

def _apply_cell_style(cell, is_number=False, is_date=False):
    cell.border = CELL_BORDER
    if is_number:
        cell.number_format = NUMBER_FORMAT
    if is_date:
        cell.number_format = DATE_FORMAT

def export_quotes_to_excel(queryset, status_labels, max_rows=1000):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Cotizaciones'
    headers = ['Número', 'Cliente', 'Vendedor', 'Estatus', 'Total', 'Moneda', 'Vigencia', 'Fecha emisión']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        _apply_header_style(cell)
    for row_idx, q in enumerate(queryset[:max_rows], 2):
        ws.cell(row=row_idx, column=1, value=q.quote_number)
        ws.cell(row=row_idx, column=2, value=q.customer.name)
        ws.cell(row=row_idx, column=3, value=q.sales_user.get_full_name() or q.sales_user.username)
        ws.cell(row=row_idx, column=4, value=status_labels.get(q.status, q.status))
        ws.cell(row=row_idx, column=5, value=float(q.total))
        ws.cell(row=row_idx, column=6, value=q.currency or '')
        ws.cell(row=row_idx, column=7, value=q.valid_until)
        ws.cell(row=row_idx, column=8, value=q.issue_date)
        for col in range(1, 9):
            cell = ws.cell(row=row_idx, column=col)
            _apply_cell_style(cell, is_number=col == 5, is_date=col in (7, 8))
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = max(12, min(25, 15))
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def export_report_to_excel(queryset, status_labels, max_rows=2000):
    return export_quotes_to_excel(queryset, status_labels, max_rows=max_rows)
